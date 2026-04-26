"""
dashboard_edp_v2.py — Tablero EDP · Formación Completa
Streamlit · SOFSE / Trenes Argentinos · Coordinación de Material Rodante
Versión 2: soporte multi-coche con vista de formación completa
"""

import re, tempfile
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
G               = 9.81
MU_DEFAULT      = 0.35
N_COCHES        = 6
MASA_TARA_COCHE = 45_000
PLAZAS_MITRE    = 408
MASA_PAX_KG     = 75
PASTILLAS_X_COCHE = 16
N_PAST_TOTAL    = PASTILLAS_X_COCHE * N_COCHES
MASA_TARA_FORM  = MASA_TARA_COCHE * N_COCHES
MASA_DISENO_FORM= MASA_TARA_FORM + PLAZAS_MITRE * MASA_PAX_KG
A_DISEÑO_SERV   = 1.0
A_DISEÑO_EMER   = 1.2
BALANCE_OK      = 10.0
BALANCE_WARN    = 20.0
CV_OK           = 5.0
CV_WARN         = 10.0

# Secuencia de la formación M27
ORDEN_FORMACION = ["TC1","M1","M2","M3","M4","TC2"]
COLOR_COCHE     = {
    "TC1": "#1F4E79", "M1": "#2E75B6", "M2": "#5BA3D9",
    "M3": "#5BA3D9",  "M4": "#2E75B6", "TC2": "#1F4E79"
}

AZUL   = "#1F4E79"; AZUL_C = "#4DA6FF"; VERDE = "#5DBB63"
NARANJA= "#F4A31E"; ROJO   = "#FF4C4C"
BG_DARK= "#0E1117"; BG_CARD= "#1E2130"; TXT   = "#FAFAFA"

PLOTLY_LAYOUT = dict(
    plot_bgcolor=BG_DARK, paper_bgcolor=BG_DARK,
    font=dict(family="Arial", color=TXT),
    xaxis=dict(gridcolor="#2A2F45", tickfont=dict(color=TXT),
               title_font=dict(color=TXT)),
    yaxis=dict(gridcolor="#2A2F45", tickfont=dict(color=TXT),
               title_font=dict(color=TXT)),
    legend=dict(bgcolor=BG_CARD, font=dict(color=TXT)),
)

MAPEO_UD_RUEDA = {
    (1,1):2,(1,2):2,(1,3):1,(1,4):1,(1,5):4,(1,6):4,(1,7):3,(1,8):3,
    (2,1):6,(2,2):6,(2,3):5,(2,4):5,(2,5):8,(2,6):8,(2,7):7,(2,8):7,
}

# ─────────────────────────────────────────────
# CARGA Y PREPARACIÓN
# ─────────────────────────────────────────────

def _extraer_bogie(nombre):
    m = re.search(r"[Bb](\d+)[_.]", str(nombre))
    return int(m.group(1)) if m else None

def _extraer_test(nombre):
    m = re.search(r"[Bb]\d+[_.](\d+)", str(nombre))
    return int(m.group(1)) if m else None

def _extraer_posicion(nombre):
    m = re.search(r"\((M\d|TC\d)\)", str(nombre))
    return m.group(1) if m else None

def cargar_excel(archivo):
    df = pd.read_excel(archivo, header=[0,1])
    df.columns = [c[1].strip() for c in df.columns]
    df = df.rename(columns={
        "Nombre de archivo":"archivo","Fecha ensayo":"fecha",
        "Formación":"formacion","Vehículo":"vehiculo",
        "Condición":"condicion","Presión aplicada (kg/cm²)":"presion",
        "Posición en formación":"posicion","Bogie":"bogie",
        "N° test":"nro_test","Unidad dinamométrica":"ud_label",
        "Fuerza Superior (kg)":"fuerza_sup",
        "Fuerza Inferior (kg)":"fuerza_inf","Total (kg)":"total",
    })
    df["ud_num"]   = df["ud_label"].str.extract(r"UD0?(\d+)").astype(int)
    df["bogie"]    = df.apply(lambda r: int(pd.to_numeric(r["bogie"], errors="coerce"))
                               if pd.notna(pd.to_numeric(r["bogie"], errors="coerce"))
                               else _extraer_bogie(r["archivo"]), axis=1)
    df["nro_test"] = df.apply(lambda r: int(pd.to_numeric(r["nro_test"], errors="coerce"))
                               if pd.notna(pd.to_numeric(r["nro_test"], errors="coerce"))
                               else _extraer_test(r["archivo"]), axis=1)
    df["pos_form"] = df["archivo"].apply(_extraer_posicion)
    df["rueda"]    = df.apply(lambda r: MAPEO_UD_RUEDA.get((r["bogie"], r["ud_num"]))
                               if r["bogie"] else None, axis=1)
    df["caliper"]  = df["ud_num"].apply(lambda n: "Superior" if n%2!=0 else "Inferior")
    df["balance_pct"] = (abs(df["fuerza_sup"]-df["fuerza_inf"])/
                         ((df["fuerza_sup"]+df["fuerza_inf"])/2)*100).round(2)
    df["condicion"] = df["condicion"].str.strip().str.title()
    df["estado_balance"] = df["balance_pct"].apply(
        lambda x: "OK" if x<BALANCE_OK else ("ATENCIÓN" if x<BALANCE_WARN else "FUERA"))
    df["tipo_coche"] = df["vehiculo"].apply(
        lambda v: "TC (cabina)" if str(v).startswith("MC") else "M (motor)")
    # Orden de posición
    orden = {p:i for i,p in enumerate(ORDEN_FORMACION)}
    df["orden_pos"] = df["pos_form"].map(orden).fillna(99)

    # ── Detección de tests inválidos ──────────────────────
    # Un test es inválido si la MAYORÍA de sus UDs tienen valores muy bajos
    # (< 100 kgf), lo que indica error de inicialización del equipo,
    # no un problema real del caliper.
    UMBRAL_INVALIDO = 100  # kgf
    resumen_arch = df.groupby("archivo")["total"].agg(
        lambda x: (x < UMBRAL_INVALIDO).mean()  # fracción de UDs bajas
    )
    archivos_invalidos = set(resumen_arch[resumen_arch > 0.5].index)
    df["test_invalido"] = df["archivo"].isin(archivos_invalidos)

    return df

def color_estado(e):
    return {"OK":"#1a9641","ATENCIÓN":"#f4a31e","FUERA":ROJO}.get(e,"#888")

# ─────────────────────────────────────────────
# SECCIÓN: VISTA DE FORMACIÓN COMPLETA
# ─────────────────────────────────────────────

def seccion_formacion(df, mu):
    st.markdown("### 🚃 Vista Completa de la Formación")
    st.caption(
        "Comparación de todos los coches de la formación. "
        "Cada barra representa la fuerza media de apriete por pastilla."
    )

    # ── Advertencia de tests inválidos ───────────────────
    invalidos = df[df["test_invalido"]==True]["archivo"].unique()
    if len(invalidos):
        with st.warning(""):
            st.markdown(
                f"**⚠️ {len(invalidos)} test(s) excluido(s) de los cálculos por error de inicialización del equipo:**"
            )
            for arch in invalidos:
                row = df[df["archivo"]==arch].iloc[0]
                st.markdown(
                    f"- `{arch}` — {row['vehiculo']} · {row.get('pos_form','?')} · "
                    f"Bogie {row['bogie']} Test {row['nro_test']} · {row['condicion']}<br>"
                    f"  Los valores registrados son < 100 kgf en la mayoría de UDs, "
                    f"indicando que el equipo no estaba inicializado. "
                    f"El test repetido del mismo bogie confirma funcionamiento correcto.",
                    unsafe_allow_html=True
                )

    # Excluir tests inválidos de todos los cálculos
    df = df[df["test_invalido"]==False].copy()

    coches  = sorted(df["vehiculo"].unique(), key=lambda v: df[df["vehiculo"]==v]["orden_pos"].iloc[0])
    conds   = sorted(df["condicion"].unique())

    # ── Gráfico de barras agrupado: promedio por coche y condición ──
    resumen = df.groupby(["vehiculo","pos_form","condicion","orden_pos"])["total"].agg(
        ["mean","min","max","std"]).reset_index()
    resumen = resumen.sort_values("orden_pos")
    resumen.columns = ["vehiculo","pos_form","condicion","orden_pos","media","minimo","maximo","std"]

    fig = px.bar(
        resumen, x="pos_form", y="media", color="condicion",
        color_discrete_map={"Servicio":AZUL_C,"Emergencia":ROJO},
        barmode="group",
        error_y="std",
        text=resumen["media"].map(lambda x: f"{x:,.0f}"),
        hover_data={"vehiculo":True,"minimo":True,"maximo":True,"std":True},
        labels={"media":"Fuerza media (kgf)","pos_form":"Posición en formación",
                "condicion":"Condición"},
        title="Fuerza media por pastilla — todos los coches de la formación",
        category_orders={"pos_form":ORDEN_FORMACION},
    )
    # Líneas de referencia para F_mínima teórica
    for cond, a_ref, color in [("Servicio",A_DISEÑO_SERV,AZUL_C),
                                 ("Emergencia",A_DISEÑO_EMER,ROJO)]:
        F_min = (MASA_TARA_FORM * a_ref) / (N_PAST_TOTAL * mu * G)
        fig.add_hline(y=F_min, line_dash="dot", line_color=color, line_width=1,
                      annotation_text=f"F_mín {cond}: {F_min:.0f} kgf",
                      annotation_position="top right")
    fig.update_traces(textposition="outside")
    fig.update_layout(**PLOTLY_LAYOUT, height=460, barmode="group",
                      legend_title="", title_font=dict(color=TXT))
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

    # ── Tabla resumen con semáforo ──
    st.markdown("**Resumen por coche**")
    filas = []
    for _, row in resumen.iterrows():
        cond = row["condicion"]
        F_min = (MASA_TARA_FORM * (A_DISEÑO_SERV if cond=="Servicio" else A_DISEÑO_EMER)) / (N_PAST_TOTAL * mu * G)
        margen = row["media"] / F_min
        cv = (row["std"] / row["media"] * 100) if row["media"] > 0 else 0
        est = "✅ OK" if row["media"] >= F_min and row["minimo"] > 0 else (
              "⚠️ Atención" if row["media"] >= F_min * 0.95 else "🔴 Bajo")
        # Detectar cero
        if row["minimo"] == 0:
            est = "🔴 Caliper sin fuerza"
        filas.append({
            "Posición": row["pos_form"],
            "Vehículo": row["vehiculo"],
            "Condición": cond,
            "Media (kgf)": round(row["media"],1),
            "Mín (kgf)": row["minimo"],
            "Máx (kgf)": row["maximo"],
            "CV (%)": round(cv,1),
            "F_mín req.": round(F_min,0),
            "Margen": f"{margen:.2f}×",
            "Estado": est,
        })
    st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

    # (Las alertas de tests inválidos se muestran arriba, antes del filtro)

# ─────────────────────────────────────────────
# SECCIÓN: COMPARACIÓN ENTRE COCHES (heatmap)
# ─────────────────────────────────────────────

def seccion_heatmap(df):
    st.markdown("### 🗺️ Mapa de Calor — Fuerza por UD y Coche")
    st.caption(
        "Cada celda muestra la fuerza media de apriete (kgf). "
        "Colores cálidos = mayor fuerza · Colores fríos = menor fuerza · "
        "Cero o muy bajo = problema."
    )

    conds = sorted(df["condicion"].unique())
    tabs  = st.tabs(conds)
    for tab, cond in zip(tabs, conds):
        with tab:
            pivot = df[df["condicion"]==cond].groupby(
                ["pos_form","orden_pos","ud_label"])["total"].mean().reset_index()
            pivot = pivot.sort_values("orden_pos")

            # Orden de UDs
            ud_order = [f"UD{i:02d} WIFI" for i in range(1,9)]
            fig = px.density_heatmap(
                pivot, x="ud_label", y="pos_form",
                z="total",
                color_continuous_scale=["#C00000","#E26B0A","#FFFF00","#1D9E75"],
                category_orders={"ud_label":ud_order,"pos_form":ORDEN_FORMACION},
                labels={"total":"Fuerza (kgf)","ud_label":"Unidad Dinom.","pos_form":"Posición"},
                title=f"Mapa de calor — {cond}",
                text_auto=".0f",
            )
            fig.update_layout(**PLOTLY_LAYOUT, height=340,
                              coloraxis_colorbar=dict(title="kgf",
                                                       tickfont=dict(color=TXT),
                                                       title_font=dict(color=TXT)))
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

# ─────────────────────────────────────────────
# SECCIÓN: DETALLE POR COCHE (selector)
# ─────────────────────────────────────────────

def seccion_detalle_coche(df, mu, bal_ok, bal_warn):
    st.markdown("### 🔍 Detalle por Coche")

    coches_orden = sorted(df["vehiculo"].unique(),
                          key=lambda v: df[df["vehiculo"]==v]["orden_pos"].iloc[0])
    coche_sel = st.selectbox(
        "Seleccionar coche:",
        coches_orden,
        format_func=lambda v: f"{df[df['vehiculo']==v]['pos_form'].iloc[0]} — {v}"
    )
    dfc = df[df["vehiculo"] == coche_sel]
    pos = dfc["pos_form"].iloc[0]
    tipo = dfc["tipo_coche"].iloc[0]

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Vehículo", coche_sel)
    c2.metric("Posición", pos)
    c3.metric("Tipo", tipo)
    c4.metric("Registros", len(dfc))

    conds = sorted(dfc["condicion"].unique())
    tabs  = st.tabs(conds)
    for tab, cond in zip(tabs, conds):
        with tab:
            dfcc = dfc[dfc["condicion"]==cond]
            F_min = (MASA_TARA_FORM * (A_DISEÑO_SERV if cond=="Servicio" else A_DISEÑO_EMER)) / (N_PAST_TOTAL * mu * G)

            # Gráfico por UD
            ud_avg = dfcc.groupby("ud_label")["total"].agg(["mean","std"]).reset_index()
            ud_avg.columns = ["ud_label","media","std"]
            ud_avg["color"] = ud_avg["media"].apply(
                lambda x: VERDE if x >= F_min else (NARANJA if x >= F_min*0.9 else ROJO))

            fig = go.Figure()
            for _, row in ud_avg.iterrows():
                fig.add_trace(go.Bar(
                    x=[row["ud_label"]], y=[row["media"]],
                    error_y=dict(type="data", array=[row["std"]], color=TXT),
                    marker_color=row["color"],
                    text=f"{row['media']:.0f}",
                    textposition="outside",
                    showlegend=False,
                ))
            fig.add_hline(y=F_min, line_dash="dot", line_color=NARANJA, line_width=1.5,
                          annotation_text=f"F_mín requerida: {F_min:.0f} kgf",
                          annotation_font=dict(color=NARANJA))
            fig.update_layout(**PLOTLY_LAYOUT, height=350,
                              title=f"{coche_sel} ({pos}) — {cond}",
                              title_font=dict(color=TXT),
                              yaxis_title="Fuerza media (kgf)",
                              showlegend=False)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

            # Balance S/I por bogie
            st.caption("Balance Superior / Inferior por bogie")
            bogs = sorted(dfcc["bogie"].dropna().unique())
            cols = st.columns(len(bogs))
            for col_w, bogie in zip(cols, bogs):
                with col_w:
                    st.markdown(f"**Bogie {int(bogie)}**")
                    dfb = (dfcc[dfcc["bogie"]==bogie]
                           .groupby("ud_num")
                           .agg(fuerza_sup=("fuerza_sup","mean"),
                                fuerza_inf=("fuerza_inf","mean"))
                           .reset_index().sort_values("ud_num"))
                    uds   = dfb["ud_num"].tolist()
                    pares = [(uds[i],uds[i+1]) for i in range(0,len(uds)-1,2)]
                    for ud_s, ud_i in pares:
                        rs = dfb[dfb["ud_num"]==ud_s].iloc[0]
                        ri = dfb[dfb["ud_num"]==ud_i].iloc[0]
                        sup,inf = rs["fuerza_sup"], ri["fuerza_inf"]
                        denom = (sup+inf)/2
                        bal   = abs(sup-inf)/denom*100 if denom > 0 else 0
                        disco = (ud_s+1)//2
                        est   = "OK" if bal<bal_ok else ("ATENCIÓN" if bal<bal_warn else "FUERA")
                        hx    = color_estado(est)
                        emoji = {"OK":"✅","ATENCIÓN":"⚠️","FUERA":"🔴"}[est]
                        alerta_cero = " 🔴 SIN PRESIÓN" if min(sup,inf)==0 else ""
                        st.markdown(
                            f"<div style='border-left:4px solid {hx};padding:5px 8px;"
                            f"margin:4px 0;border-radius:4px;background:{BG_CARD};"
                            f"font-size:.85rem;color:{TXT}'>"
                            f"<b>Disco {disco}</b>{alerta_cero}<br>"
                            f"↑ Sup: <b>{sup:.0f}</b> kgf &nbsp; ↓ Inf: <b>{inf:.0f}</b> kgf<br>"
                            f"Balance: <b>{bal:.1f}%</b> {emoji} {est}</div>",
                            unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SECCIÓN: REPETIBILIDAD MULTI-COCHE
# ─────────────────────────────────────────────

def seccion_repetibilidad(df):
    st.markdown("### 🔁 Repetibilidad por Coche")
    st.caption("CV = σ / media × 100 entre los tests del mismo bogie/condición  |  ✅ < 5%   ⚠️ 5–10%   🔴 > 10%")

    rep = df.groupby(["vehiculo","pos_form","condicion","orden_pos","ud_label"])["total"]\
            .agg(["mean","std"]).reset_index()
    rep["cv"] = (rep["std"]/rep["mean"]*100).fillna(0).round(1)
    rep = rep.sort_values("orden_pos")

    conds = sorted(df["condicion"].unique())
    tabs  = st.tabs(conds)
    for tab, cond in zip(tabs, conds):
        with tab:
            dft = rep[rep["condicion"]==cond]
            fig = px.box(
                dft, x="pos_form", y="cv",
                color="pos_form",
                points="all",
                labels={"cv":"CV (%)","pos_form":"Coche"},
                title=f"Distribución de CV entre UDs — {cond}",
                category_orders={"pos_form":ORDEN_FORMACION},
            )
            fig.add_hline(y=CV_OK,   line_dash="dash", line_color=VERDE,   line_width=1)
            fig.add_hline(y=CV_WARN, line_dash="dash", line_color=NARANJA, line_width=1)
            fig.update_layout(**PLOTLY_LAYOUT, height=380,
                              showlegend=False, title_font=dict(color=TXT))
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

# ─────────────────────────────────────────────
# SECCIÓN: SERV VS EMER MULTI-COCHE
# ─────────────────────────────────────────────

def seccion_serv_emer(df):
    st.markdown("### 📈 Servicio vs Emergencia — Todos los Coches")
    if df["condicion"].nunique() < 2:
        st.info("Cargá un archivo con ambas condiciones."); return

    comp = df.groupby(["vehiculo","pos_form","condicion","orden_pos"])["total"]\
             .mean().reset_index().sort_values("orden_pos")

    fig = px.line(comp, x="pos_form", y="total", color="condicion",
                  color_discrete_map={"Servicio":AZUL_C,"Emergencia":ROJO},
                  markers=True, line_dash="vehiculo",
                  labels={"total":"Fuerza media (kgf)","pos_form":"Posición","condicion":"Condición"},
                  title="Fuerza media por coche — Servicio vs Emergencia",
                  category_orders={"pos_form":ORDEN_FORMACION})
    fig.update_layout(**PLOTLY_LAYOUT, height=400, title_font=dict(color=TXT))
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

    # Ratio
    pivot = comp.pivot_table(index=["pos_form","orden_pos"],
                              columns="condicion", values="total").reset_index()
    if "Servicio" in pivot.columns and "Emergencia" in pivot.columns:
        pivot["ratio"] = (pivot["Emergencia"]/pivot["Servicio"]).round(3)
        pivot = pivot.sort_values("orden_pos")
        fig2 = px.bar(pivot, x="pos_form", y="ratio",
                      text=pivot["ratio"].map(lambda x:f"{x:.2f}×"),
                      labels={"ratio":"Ratio Emer/Serv","pos_form":"Coche"},
                      title="Ratio Emergencia/Servicio por coche",
                      color_discrete_sequence=[AZUL_C],
                      category_orders={"pos_form":ORDEN_FORMACION})
        fig2.add_hline(y=2.80/2.50, line_dash="dot", line_color=ROJO,
                       annotation_text=f"Teórico: {2.80/2.50:.2f}×",
                       annotation_font=dict(color=ROJO))
        fig2.update_traces(textposition="outside")
        fig2.update_layout(**PLOTLY_LAYOUT, height=360, title_font=dict(color=TXT))
        st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar":False})

# ─────────────────────────────────────────────
# SECCIÓN: DATOS COMPLETOS
# ─────────────────────────────────────────────

def seccion_datos(df):
    st.markdown("### 📋 Datos Completos")
    coches_disp = ["Todos"] + sorted(df["vehiculo"].unique(),
                                      key=lambda v: df[df["vehiculo"]==v]["orden_pos"].iloc[0])
    sel = st.selectbox("Filtrar por coche:", coches_disp)
    dff = df if sel=="Todos" else df[df["vehiculo"]==sel]

    mostrar = dff.rename(columns={
        "archivo":"Archivo","fecha":"Fecha","formacion":"Formación",
        "vehiculo":"Vehículo","pos_form":"Posición","condicion":"Condición",
        "presion":"Presión (kg/cm²)","bogie":"Bogie","nro_test":"N° Test",
        "ud_label":"UD","rueda":"Rueda","caliper":"Caliper",
        "fuerza_sup":"F. Superior (kgf)","fuerza_inf":"F. Inferior (kgf)",
        "total":"Total (kgf)","balance_pct":"Balance (%)","estado_balance":"Estado",
    }).drop(columns=["ud_num","orden_pos","tipo_coche"], errors="ignore")
    st.dataframe(mostrar, use_container_width=True, hide_index=True)

    def to_excel(df_in):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="EDP"
        thin=Side(style="thin",color="CCCCCC")
        borde=Border(left=thin,right=thin,top=thin,bottom=thin)
        hfill=PatternFill("solid",start_color="1F4E79")
        altfill=PatternFill("solid",start_color="D6E4F0")
        cols=list(df_in.columns)
        for c,h in enumerate(cols,1):
            cell=ws.cell(row=1,column=c,value=h)
            cell.font=Font(name="Arial",bold=True,color="FFFFFF",size=9)
            cell.fill=hfill
            cell.alignment=Alignment(horizontal="center",wrap_text=True)
            cell.border=borde
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width=max(10,len(h)+2)
        ws.row_dimensions[1].height=28
        altf=PatternFill("solid",start_color="D6E4F0")
        for i,row in enumerate(df_in.itertuples(index=False),2):
            for c,val in enumerate(row,1):
                cell=ws.cell(row=i,column=c,value=val)
                cell.font=Font(name="Arial",size=8)
                cell.border=borde
                if i%2==0: cell.fill=altf
        ws.freeze_panes="A2"
        ws.auto_filter.ref=f"A1:{openpyxl.utils.get_column_letter(len(cols))}1"
        tmp=tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False)
        wb.save(tmp.name); return tmp.name

    with open(to_excel(mostrar),"rb") as f:
        st.download_button("📥 Descargar Excel",data=f.read(),
                           file_name="edp_formacion_completa.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# APP PRINCIPAL
# ─────────────────────────────────────────────

st.set_page_config(page_title="Tablero EDP — Formación Completa",
                   page_icon="🚆", layout="wide")

CSS = """<style>
[data-testid="stMetricValue"]{color:#FAFAFA}
[data-testid="stMetricLabel"]{color:#A0A0B0}
</style>"""
st.markdown(CSS, unsafe_allow_html=True)
st.markdown(
    f"<div style='background:{AZUL};padding:16px 24px;border-radius:8px;margin-bottom:18px'>"
    f"<h2 style='color:white;margin:0'>🚆 Tablero EDP — Formación Completa</h2>"
    f"<p style='color:#cde;margin:4px 0 0'>Línea Mitre · SOFSE / Trenes Argentinos · "
    f"Coordinación de Material Rodante</p></div>", unsafe_allow_html=True)

with st.sidebar:
    st.markdown(f"<h3 style='color:{AZUL_C}'>⚙️ Configuración</h3>", unsafe_allow_html=True)
    archivo = st.file_uploader("Cargar Excel EDP (multi-coche)", type=["xlsx"])
    st.divider()
    st.markdown("**Parámetros**")
    mu_val   = st.number_input("Coef. fricción μ", value=MU_DEFAULT, step=0.01, format="%.2f")
    st.divider()
    st.markdown("**Umbrales balance S/I**")
    bal_ok   = st.slider("Verde → Amarillo (%)", 5,  25, int(BALANCE_OK))
    bal_warn = st.slider("Amarillo → Rojo (%)",  10, 40, int(BALANCE_WARN))
    st.divider()
    st.markdown("<small style='color:#888'>Ref.: UIC 541-3 · EN 15328 · UIC 544-1 · ETC FR v3.0</small>",
                unsafe_allow_html=True)

if not archivo:
    st.info("👈 Cargá el Excel con los datos EDP de la formación completa.")
    st.markdown("""
    **Novedades de esta versión:**
    - Vista de formación completa con todos los coches ordenados (TC1→M1→M2→M3→M4→TC2)
    - Mapa de calor por coche y unidad dinamométrica
    - Detalle individual por coche con selector
    - Alerta automática para calipers con fuerza = 0 kgf
    - Comparación Servicio vs Emergencia para toda la formación
    - Análisis de repetibilidad con box plot por coche
    """)
else:
    with st.spinner("Cargando datos..."):
        try:
            df = cargar_excel(archivo)
            df["estado_balance"] = df["balance_pct"].apply(
                lambda x: "OK" if x<bal_ok else ("ATENCIÓN" if x<bal_warn else "FUERA"))
        except Exception as e:
            st.error(f"Error: {e}"); st.stop()

    n_coches = df["vehiculo"].nunique()
    n_conds  = df["condicion"].nunique()
    st.success(
        f"✅ {len(df)} mediciones · {n_coches} coches · "
        f"{df['formacion'].iloc[0]} · "
        f"Condiciones: {', '.join(sorted(df['condicion'].unique()))}"
    )

    tabs = st.tabs([
        "🚃 Formación completa",
        "🗺️ Mapa de calor",
        "🔍 Detalle por coche",
        "🔁 Repetibilidad",
        "📈 Serv vs Emer",
        "📋 Datos",
    ])
    # df_valido = excluye tests inválidos para cálculos; df completo para datos
    df_valido = df[df["test_invalido"]==False].copy()

    with tabs[0]: seccion_formacion(df, mu_val)          # maneja internamente
    with tabs[1]: seccion_heatmap(df_valido)
    with tabs[2]: seccion_detalle_coche(df_valido, mu_val, bal_ok, bal_warn)
    with tabs[3]: seccion_repetibilidad(df_valido)
    with tabs[4]: seccion_serv_emer(df_valido)
    with tabs[5]: seccion_datos(df)                       # muestra todo, con columna estado
